import os
import json
import re
import io
import csv
import sqlite3
import subprocess
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, url_for, send_file
from werkzeug.utils import secure_filename

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'inventario.db')
PUBLIC_DIR = os.path.join(BASE_DIR, 'public')
IMAGES_DIR = os.path.join(PUBLIC_DIR, 'imagenes')
JSON_CATALOG_PATH = os.path.join(PUBLIC_DIR, 'productos.json')

def sincronizar_con_github():
    """Ejecuta los comandos de Git en un hilo secundario sin congelar la interfaz ni retrasar las peticiones del usuario."""
    def tarea():
        try:
            # 1. Asegurar que git rastree el JSON actualizado y todas las imágenes nuevas
            subprocess.run(["git", "add", "public/productos.json", "public/imagenes/"], cwd=BASE_DIR, check=True)
            
            # 2. Verificar si hay cambios pendientes antes de commitear
            status = subprocess.run(["git", "status", "--porcelain"], cwd=BASE_DIR, capture_output=True, text=True)
            if status.stdout.strip():
                subprocess.run(["git", "commit", "-m", "Auto-sync: inventario, catalogo e imagenes actualizadas"], cwd=BASE_DIR, check=True)
                subprocess.run(["git", "push"], cwd=BASE_DIR, check=True)
                print("🚀 Cambios sincronizados exitosamente con GitHub y Vercel.")
            else:
                print("ℹ️ No hay cambios pendientes para sincronizar con Git.")
        except Exception as e:
            print(f"⚠️ Error en la sincronización automática con Git: {e}")

    threading.Thread(target=tarea, daemon=True).start()

def normalizar_nombre_imagen(filename):
    """
    Normaliza el nombre de archivo de imagen asegurando que la extensión esté en minúsculas
    y coincida exactamente con el archivo real en disco (public/imagenes/) para evitar discrepancias .JPG vs .jpg.
    """
    if not filename:
        return ""
    filename_str = str(filename).strip()
    if not filename_str or filename_str.startswith('http://') or filename_str.startswith('https://') or filename_str.startswith('data:'):
        return filename_str

    clean_name = filename_str.replace('\\', '/')
    prefix = ""
    if clean_name.startswith('public/imagenes/'):
        prefix = "imagenes/"
        clean_name = clean_name[16:]
    elif clean_name.startswith('imagenes/'):
        prefix = "imagenes/"
        clean_name = clean_name[9:]

    base, ext = os.path.splitext(clean_name)
    ext_lower = ext.lower()
    normalized_file = f"{base}{ext_lower}"

    if os.path.exists(IMAGES_DIR):
        try:
            files_on_disk = os.listdir(IMAGES_DIR)
            for disk_file in files_on_disk:
                if disk_file.lower() == normalized_file.lower():
                    normalized_file = disk_file
                    break
        except Exception:
            pass

    return f"{prefix}{normalized_file}" if prefix else normalized_file

os.makedirs(IMAGES_DIR, exist_ok=True)

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'), static_folder=PUBLIC_DIR)
app.config['UPLOAD_FOLDER'] = IMAGES_DIR
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16 MB max

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def parse_variantes_string(var_str):
    """
    Parsea una cadena de variantes (ej: "M:10, L:8, XL:6" o "Negro:112, Rojo:34" o "M, L, XL")
    a una lista de objetos [{"nombre": "M", "stock": 10}, ...]
    """
    if not var_str or not str(var_str).strip():
        return []
    
    parts = [p.strip() for p in str(var_str).split(',') if p.strip()]
    variantes_list = []
    
    for part in parts:
        if ':' in part:
            name, stock_str = part.split(':', 1)
            name = name.strip()
            try:
                stock = int(stock_str.strip())
            except ValueError:
                stock = 0
            if name:
                variantes_list.append({"nombre": name, "stock": stock})
        else:
            name = part.strip()
            if name:
                variantes_list.append({"nombre": name, "stock": 0})
            
    return variantes_list

def merge_variantes_preserving_stock(existing_var_str, submitted_var_str, current_stock_actual=0):
    """
    Combina las variantes enviadas desde la interfaz web con las existentes en la BD.
    Preserva el stock de cada variante existente y el stock total del producto.
    Bajo NINGUNA circunstancia borra o resetea a 0 el stock_actual del producto.
    """
    existing_parsed = parse_variantes_string(existing_var_str)
    
    # 1. Si no hay variantes enviadas
    if not submitted_var_str or not str(submitted_var_str).strip():
        if existing_parsed and any(v['stock'] > 0 for v in existing_parsed):
            total = sum(v['stock'] for v in existing_parsed)
            return (existing_var_str or ""), max(total, current_stock_actual)
        else:
            return (existing_var_str or ""), current_stock_actual

    # 2. Si se enviaron variantes desde la web
    stock_dict = { v['nombre'].strip().upper(): v['stock'] for v in existing_parsed }
    original_name_map = { v['nombre'].strip().upper(): v['nombre'].strip() for v in existing_parsed }

    submitted_parts = [p.strip() for p in str(submitted_var_str).split(',') if p.strip()]
    merged_list = []
    seen_keys = set()
    has_explicit_stock_in_submission = False

    for part in submitted_parts:
        if ':' in part:
            name, st_str = part.split(':', 1)
            name = name.strip()
            key = name.upper()
            try:
                st = int(st_str.strip())
                has_explicit_stock_in_submission = True
            except ValueError:
                st = stock_dict.get(key, 0)
        else:
            name = part.strip()
            key = name.upper()
            st = stock_dict.get(key, 0)

        if key and key not in seen_keys:
            seen_keys.add(key)
            final_name = original_name_map.get(key, name)
            merged_list.append({"nombre": final_name, "stock": max(0, st)})

    sum_from_merged = sum(v['stock'] for v in merged_list)
    
    if sum_from_merged > 0 or has_explicit_stock_in_submission:
        new_total_stock = max(sum_from_merged, current_stock_actual)
        new_str = ", ".join([f"{v['nombre']}:{v['stock']}" for v in merged_list])
    else:
        new_total_stock = current_stock_actual
        new_str = ", ".join([v['nombre'] for v in merged_list])

    return new_str, new_total_stock

def init_db():
    conn = get_db()
    cursor = conn.cursor()
    
    # Tabla productos
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS productos (
            sku TEXT PRIMARY KEY,
            categoria TEXT NOT NULL,
            marca TEXT NOT NULL,
            descripcion TEXT NOT NULL,
            variantes TEXT,
            costo_compra REAL NOT NULL,
            precio_venta REAL NOT NULL,
            stock_actual INTEGER NOT NULL DEFAULT 0,
            stock_minimo INTEGER NOT NULL DEFAULT 2,
            imagen TEXT,
            imagenes_extra TEXT DEFAULT '',
            descripcion_detallada TEXT DEFAULT '',
            es_popular INTEGER DEFAULT 0
        )
    ''')
    
    # Tabla ventas
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS ventas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fecha_hora DATETIME NOT NULL,
            cliente TEXT NOT NULL,
            total REAL NOT NULL,
            detalle_json TEXT NOT NULL
        )
    ''')
    
    conn.commit()

    # Migración de columnas si la tabla ya existía
    cursor.execute("PRAGMA table_info(productos)")
    existing_cols = [col['name'] for col in cursor.fetchall()]
    if 'imagenes_extra' not in existing_cols:
        cursor.execute("ALTER TABLE productos ADD COLUMN imagenes_extra TEXT DEFAULT ''")
    if 'descripcion_detallada' not in existing_cols:
        cursor.execute("ALTER TABLE productos ADD COLUMN descripcion_detallada TEXT DEFAULT ''")
    if 'es_popular' not in existing_cols:
        cursor.execute("ALTER TABLE productos ADD COLUMN es_popular INTEGER DEFAULT 0")

    # Asegurar columnas necesarias en SQLite requeridas para el catálogo web
    columnas_necesarias = [
        ("imagenes_galeria", "TEXT DEFAULT '[]'"),
        ("descripcion_larga", "TEXT DEFAULT ''"),
        ("lista_variantes", "TEXT DEFAULT ''"),
        ("destacado_popular", "INTEGER DEFAULT 0"),
        ("ocultar_web", "INTEGER DEFAULT 0")
    ]
    for col_nombre, col_tipo in columnas_necesarias:
        try:
            cursor.execute(f"ALTER TABLE productos ADD COLUMN {col_nombre} {col_tipo}")
            conn.commit()
        except Exception:
            pass # La columna ya existe

    if 'imagenes_galeria' not in existing_cols:
        cursor.execute("UPDATE productos SET imagenes_galeria = imagenes_extra WHERE imagenes_galeria IS NULL OR imagenes_galeria = ''")
    if 'descripcion_larga' not in existing_cols:
        cursor.execute("UPDATE productos SET descripcion_larga = descripcion_detallada WHERE descripcion_larga IS NULL OR descripcion_larga = ''")
    if 'lista_variantes' not in existing_cols:
        cursor.execute("UPDATE productos SET lista_variantes = variantes WHERE lista_variantes IS NULL OR lista_variantes = ''")
    if 'destacado_popular' not in existing_cols:
        cursor.execute("UPDATE productos SET destacado_popular = es_popular WHERE destacado_popular IS NULL OR destacado_popular = 0")
    conn.commit()

    # Seeding inicial si la tabla está vacía
    cursor.execute('SELECT COUNT(*) as count FROM productos')
    row = cursor.fetchone()
    if row['count'] == 0:
        seed_data = [
            ('1060', 'Cascos', 'Scoyco', 'Casco Integral Scoyco R-1 Certificado DOT', 'M:10, L:8, XL:6', 95.00, 145.00, 24, 3, '1060.jpg', '', 'Casco aerodinámico certificado DOT con visor anti-rayaduras y almohadillas lavables.', 1),
            ('SX20', 'Luces', 'Kaifa', 'Faro LED Auxiliar Explorer 40W Dual Beam', 'Blanca:14, Amarilla:10', 35.00, 65.00, 24, 5, 'SX20.jpg', '', 'Potencia de 40W con cambio de luz blanca y exploración amarilla para niebla.', 1),
            ('A-010', 'Accesorios', 'AGR', 'Guantes Cuero con Protección de Nudillos', 'Negro L:10, Negro XL:8, Rojo M:6, Rojo L:6', 28.00, 48.00, 30, 4, 'A-010.jpg', '', 'Guantes de alta resistencia con protección rígida en nudillos y agarre antideslizante.', 1),
            ('MQ-20', 'Aceites', 'Motul', 'Aceite Sintético Motul 7100 10W40 4T 1L', '1 Litro:40', 42.00, 62.00, 40, 8, 'MQ-20.jpg', '', 'Lubricante 100% sintético con tecnología de éster para motores 4 tiempos.', 1),
            ('K-400', 'Repuestos', 'AGR', 'Kit de Arrastre Reforzado Cadena 520H Heavy Duty', 'Paso 520:8', 75.00, 120.00, 8, 2, 'K-400.jpg', '', 'Kit de arrastre reforzado para alta cilindrada con catalina de acero al carbono.', 0),
            ('SEC-90', 'Seguridad', 'Icon', 'Alarma Antirrobo Sensor de Movimiento con Control', 'Estándar 12V:12', 45.00, 85.00, 12, 3, 'SEC-90.jpg', '', 'Alarma inteligente con sirena de 125dB y corte de encendido a distancia.', 0),
            ('ESP-05', 'Accesorios', 'Kaifa', 'Espejos Deportivos Aluminio CNC Universal', 'Negro:10, Azul:8', 22.00, 42.00, 18, 3, 'ESP-05.jpg', '', 'Espejos antirreflejo mecanizados en aleación CNC para timón universal de moto.', 1),
            ('SLI-02', 'Seguridad', 'AGR', 'Sliders Anticaída para Moto Chasis Universal', 'Rojo:0, Negro:1, Dorado:0', 30.00, 55.00, 1, 3, 'SLI-02.jpg', '', 'Protectores de chasis y carenado en teflón de alta resistencia a caídas.', 0),
            ('CH-100', 'Seguridad', 'AGR', 'Chaleco Reflectivo Alta Visibilidad Homologado', 'Neón L:0, Neón XL:0', 12.00, 25.00, 0, 5, 'CH-100.jpg', '', 'Chaleco reflectivo con bandas reflectantes de 360 grados homologado.', 0)
        ]
        cursor.executemany('''
            INSERT INTO productos (sku, categoria, marca, descripcion, variantes, costo_compra, precio_venta, stock_actual, stock_minimo, imagen, imagenes_extra, descripcion_detallada, es_popular)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', seed_data)
        cursor.execute("UPDATE productos SET imagenes_galeria = imagenes_extra, descripcion_larga = descripcion_detallada, lista_variantes = variantes, destacado_popular = es_popular")
        conn.commit()
    else:
        cursor.execute("UPDATE productos SET es_popular = 1, destacado_popular = 1 WHERE sku IN ('1060', 'SX20', 'A-010', 'MQ-20', 'ESP-05') AND (es_popular IS NULL OR es_popular = 0)")
        conn.commit()

    conn.close()
    exportar_productos_json()

def exportar_productos_json():
    """Genera automáticamente el archivo public/productos.json con la estructura exacta requerida."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT sku, categoria, marca, descripcion, variantes, lista_variantes, 
                   precio_venta, stock_actual, imagen, imagenes_extra, imagenes_galeria, 
                   descripcion_detallada, descripcion_larga, es_popular, destacado_popular,
                   ocultar_web
            FROM productos
            WHERE (ocultar_web IS NULL OR ocultar_web = 0)
            ORDER BY categoria, marca, descripcion
        ''')
        rows = cursor.fetchall()
        productos = []
        for r in rows:
            # Procesar lista de imágenes
            imgs_raw = []
            main_img = normalizar_nombre_imagen((r['imagen'] or "").strip())
            if main_img:
                imgs_raw.append(main_img)
            
            extras_raw = (r['imagenes_galeria'] if 'imagenes_galeria' in r.keys() and r['imagenes_galeria'] else r['imagenes_extra']) or ""
            if extras_raw:
                try:
                    parsed_extras = json.loads(extras_raw)
                    if isinstance(parsed_extras, list):
                        for img_item in parsed_extras:
                            img_item = normalizar_nombre_imagen(str(img_item).strip())
                            if img_item and img_item not in imgs_raw:
                                imgs_raw.append(img_item)
                except Exception:
                    for img_item in str(extras_raw).split(','):
                        img_item = normalizar_nombre_imagen(img_item.strip())
                        if img_item and img_item not in imgs_raw:
                            imgs_raw.append(img_item)
            
            # Formatear rutas a "imagenes/..." según estándar del json
            formatted_imgs = []
            for img in imgs_raw:
                if not img:
                    continue
                if img.startswith('http') or img.startswith('imagenes/'):
                    formatted_imgs.append(img)
                else:
                    formatted_imgs.append(f"imagenes/{img}")

            var_str = (r['lista_variantes'] if 'lista_variantes' in r.keys() and r['lista_variantes'] else r['variantes']) or ""
            var_parsed = parse_variantes_string(var_str)

            desc_text = (r['descripcion_larga'] if 'descripcion_larga' in r.keys() and r['descripcion_larga'] else r['descripcion_detallada']) or r['descripcion']

            is_popular = bool(r['destacado_popular']) if ('destacado_popular' in r.keys() and r['destacado_popular'] is not None) else (bool(r['es_popular']) if r['es_popular'] is not None else False)

            stock_actual_db = int(r['stock_actual'])
            if var_parsed and any(':' in part for part in str(var_str).split(',')):
                stock_total_val = sum(v['stock'] for v in var_parsed)
            else:
                stock_total_val = stock_actual_db

            productos.append({
                "sku": r['sku'],
                "nombre": r['descripcion'],
                "marca": r['marca'],
                "categoria": r['categoria'],
                "precio": float(r['precio_venta']),
                "stock_total": stock_total_val,
                "variantes": var_parsed,
                "imagenes": formatted_imgs if len(formatted_imgs) > 0 else (["imagenes/" + main_img] if main_img else []),
                "descripcion": desc_text,
                "popular": is_popular,
                
                # Campos de retrocompatibilidad
                "stock": stock_total_val,
                "stock_actual": stock_total_val,
                "precio_venta": float(r['precio_venta']),
                "descripcion_detallada": desc_text,
                "imagen": main_img,
                "es_popular": is_popular,
                "variantes_raw": var_str
            })
        conn.close()
        
        with open(JSON_CATALOG_PATH, 'w', encoding='utf-8') as f:
            json.dump(productos, f, ensure_ascii=False, indent=2)
        
        # Disparar auto-sincronización con GitHub y Vercel en segundo plano
        sincronizar_con_github()
        return True
    except Exception as e:
        print(f"Error al exportar productos.json: {e}")
        return False

# Inicializar DB al arrancar el modulo
init_db()

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/catalogo')
def public_catalog():
    return send_from_directory(PUBLIC_DIR, 'index.html')

@app.route('/imagenes/<path:filename>')
@app.route('/public/imagenes/<path:filename>')
def serve_images(filename):
    return send_from_directory(IMAGES_DIR, filename)

@app.route('/productos.json')
@app.route('/public/productos.json')
def serve_productos_json():
    return send_from_directory(PUBLIC_DIR, 'productos.json')

@app.route('/public/<path:filename>')
def serve_public_files(filename):
    return send_from_directory(PUBLIC_DIR, filename)

# API ENDPOINTS

@app.route('/api/productos', methods=['GET'])
def get_productos():
    query = request.args.get('q', '').strip()
    status_filter = request.args.get('status', 'all').strip().lower()
    
    conn = get_db()
    cursor = conn.cursor()
    
    sql = 'SELECT * FROM productos WHERE 1=1'
    params = []
    
    if query:
        sql += ' AND (sku LIKE ? OR marca LIKE ? OR categoria LIKE ? OR descripcion LIKE ?)'
        pattern = f'%{query}%'
        params.extend([pattern, pattern, pattern, pattern])
        
    if status_filter == 'instock':
        sql += ' AND stock_actual > stock_minimo'
    elif status_filter == 'reponer':
        sql += ' AND stock_actual > 0 AND stock_actual <= stock_minimo'
    elif status_filter == 'agotado':
        sql += ' AND stock_actual = 0'
        
    sql += ' ORDER BY sku ASC'
    
    cursor.execute(sql, params)
    rows = cursor.fetchall()
    conn.close()
    
    productos = []
    for r in rows:
        keys = r.keys()
        
        sku = r['sku']
        cat = r['categoria']
        marca = r['marca']
        desc = r['descripcion']
        var_raw = r['variantes'] if 'variantes' in keys else ''
        lista_var_raw = r['lista_variantes'] if ('lista_variantes' in keys and r['lista_variantes']) else var_raw
        
        costo = float(r['costo_compra']) if ('costo_compra' in keys and r['costo_compra'] is not None) else 0.0
        precio = float(r['precio_venta']) if ('precio_venta' in keys and r['precio_venta'] is not None) else 0.0
        stock = int(r['stock_actual']) if ('stock_actual' in keys and r['stock_actual'] is not None) else 0
        stock_min = int(r['stock_minimo']) if ('stock_minimo' in keys and r['stock_minimo'] is not None) else 2
        
        ganancia_unidad = precio - costo
        markup_pct = ((precio - costo) / costo * 100) if costo > 0 else 0
        margen_pct = ((precio - costo) / precio * 100) if precio > 0 else 0

        if stock == 0:
            estado_code = "agotado"
            estado_label = "Agotado"
            estado_class = "bg-rose-100 text-rose-800 border-rose-200"
        elif stock <= stock_min:
            estado_code = "reponer"
            estado_label = "Reponer"
            estado_class = "bg-amber-100 text-amber-800 border-amber-200"
        else:
            estado_code = "instock"
            estado_label = "En Stock"
            estado_class = "bg-emerald-100 text-emerald-800 border-emerald-200"

        imagen = r['imagen'] if ('imagen' in keys and r['imagen']) else f"{sku}.jpg"
        raw_galeria = r['imagenes_galeria'] if ('imagenes_galeria' in keys and r['imagenes_galeria']) else (r['imagenes_extra'] if ('imagenes_extra' in keys and r['imagenes_extra']) else '')
        raw_desc_larga = r['descripcion_larga'] if ('descripcion_larga' in keys and r['descripcion_larga']) else (r['descripcion_detallada'] if ('descripcion_detallada' in keys and r['descripcion_detallada']) else desc)
        pop_raw = r['destacado_popular'] if ('destacado_popular' in keys and r['destacado_popular'] is not None) else (r['es_popular'] if ('es_popular' in keys and r['es_popular'] is not None) else 0)
        ocultar_web = bool(r['ocultar_web']) if ('ocultar_web' in keys and r['ocultar_web'] is not None) else False
        
        productos.append({
            'sku': sku,
            'categoria': cat,
            'marca': marca,
            'descripcion': desc,
            'variantes': lista_var_raw,
            'lista_variantes': lista_var_raw,
            'variantes_raw': lista_var_raw,
            'costo_compra': costo,
            'precio_venta': precio,
            'stock_actual': stock,
            'stock_minimo': stock_min,
            'estado': estado_code,
            'estado_label': estado_label,
            'estado_class': estado_class,
            'imagen': imagen,
            'imagenes_extra': raw_galeria,
            'imagenes_galeria': raw_galeria,
            'descripcion_detallada': raw_desc_larga,
            'descripcion_larga': raw_desc_larga,
            'es_popular': bool(pop_raw in [1, True, '1', 'true', 'True']),
            'destacado_popular': bool(pop_raw in [1, True, '1', 'true', 'True']),
            'ocultar_web': ocultar_web,
            'ganancia_unidad': round(ganancia_unidad, 2),
            'markup_pct': round(markup_pct, 1),
            'margen_pct': round(margen_pct, 1)
        })
        
    return jsonify(productos)

@app.route('/api/productos/<sku>', methods=['GET'])
@app.route('/api/producto/<sku>', methods=['GET'])
def get_producto(sku):
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM productos WHERE UPPER(sku) = ?', (sku.strip().upper(),))
        r = cursor.fetchone()
        conn.close()
        
        if not r:
            return jsonify({'error': f'Producto con SKU "{sku}" no encontrado'}), 404
            
        keys = r.keys()
        
        costo = float(r['costo_compra']) if ('costo_compra' in keys and r['costo_compra'] is not None) else 0.0
        precio = float(r['precio_venta']) if ('precio_venta' in keys and r['precio_venta'] is not None) else 0.0
        stock_act = int(r['stock_actual']) if ('stock_actual' in keys and r['stock_actual'] is not None) else 0
        stock_min = int(r['stock_minimo']) if ('stock_minimo' in keys and r['stock_minimo'] is not None) else 0
        
        desc = r['descripcion'] if 'descripcion' in keys else ''
        marca = r['marca'] if 'marca' in keys else ''
        categoria = r['categoria'] if 'categoria' in keys else ''
        imagen = r['imagen'] if 'imagen' in keys else ''
        
        raw_galeria = r['imagenes_galeria'] if ('imagenes_galeria' in keys and r['imagenes_galeria']) else (r['imagenes_extra'] if ('imagenes_extra' in keys and r['imagenes_extra']) else '')
        raw_desc_larga = r['descripcion_larga'] if ('descripcion_larga' in keys and r['descripcion_larga']) else (r['descripcion_detallada'] if ('descripcion_detallada' in keys and r['descripcion_detallada']) else desc)
        raw_variantes = r['lista_variantes'] if ('lista_variantes' in keys and r['lista_variantes']) else (r['variantes'] if ('variantes' in keys and r['variantes']) else '')
        
        pop_raw = r['destacado_popular'] if ('destacado_popular' in keys and r['destacado_popular'] is not None) else (r['es_popular'] if ('es_popular' in keys and r['es_popular'] is not None) else 0)
        is_pop = 1 if pop_raw in [1, True, '1', 'true', 'True'] else 0
        ocultar_web_val = bool(r['ocultar_web']) if ('ocultar_web' in keys and r['ocultar_web'] is not None) else False

        return jsonify({
            'sku': r['sku'],
            'nombre': desc,
            'marca': marca,
            'categoria': categoria,
            'descripcion': desc,
            'costo_compra': costo,
            'precio_venta': precio,
            'stock_actual': stock_act,
            'stock_minimo': stock_min,
            'imagen': imagen or '',
            'imagenes_extra': raw_galeria or '',
            'imagenes_galeria': raw_galeria or '',
            'descripcion_detallada': raw_desc_larga or '',
            'descripcion_larga': raw_desc_larga or '',
            'variantes': raw_variantes or '',
            'lista_variantes': raw_variantes or '',
            'es_popular': bool(is_pop),
            'destacado_popular': is_pop,
            'ocultar_web': ocultar_web_val
        })
    except Exception as e:
        print(f"Error en get_producto({sku}):", e)
        return jsonify({'error': str(e)}), 500

@app.route('/api/productos/<sku>/web-config', methods=['PUT', 'PATCH'])
def update_web_config(sku):
    data = request.get_json() or {}
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku, stock_actual, variantes, lista_variantes FROM productos WHERE UPPER(sku) = ?', (sku.upper(),))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': f'El producto con SKU "{sku}" no existe'}), 404
        
    imagen_principal = normalizar_nombre_imagen(str(data.get('imagen', '') or data.get('imagen_principal', '')).strip())
    
    raw_extras = data.get('imagenes_extra') if data.get('imagenes_extra') is not None else data.get('imagenes_galeria')
    if raw_extras is None:
        raw_extras = data.get('imagenes', [])
        if isinstance(raw_extras, list) and len(raw_extras) > 0 and raw_extras[0] == imagen_principal:
            raw_extras = raw_extras[1:]

    if isinstance(raw_extras, list):
        imagenes_extra_json = json.dumps([img.strip() for img in raw_extras if img and img.strip()])
    else:
        imagenes_extra_json = str(raw_extras).strip()

    descripcion_detallada = str(data.get('descripcion_larga', '') or data.get('descripcion_detallada', '') or data.get('descripcion', '')).strip()
    
    existing_var_str = (row['lista_variantes'] if 'lista_variantes' in row.keys() and row['lista_variantes'] else row['variantes']) or ""
    submitted_var_str = str(data.get('lista_variantes', '') or data.get('variantes', '')).strip()
    current_stock = int(row['stock_actual']) if ('stock_actual' in row.keys() and row['stock_actual'] is not None) else 0

    merged_var_str, merged_stock_total = merge_variantes_preserving_stock(existing_var_str, submitted_var_str, current_stock)
    
    popular_val = data.get('destacado_popular')
    if popular_val is None:
        popular_val = data.get('es_popular')
    if popular_val is None:
        popular_val = data.get('popular')

    es_popular = 1 if popular_val in [True, 1, '1', 'true', 'True'] else 0

    ocultar_val = data.get('ocultar_web')
    if ocultar_val is None:
        ocultar_val = data.get('ocultar_en_web')
    ocultar_web = 1 if ocultar_val in [True, 1, '1', 'true', 'True'] else 0

    cursor.execute('''
        UPDATE productos
        SET imagen = ?, 
            imagenes_extra = ?, imagenes_galeria = ?, 
            descripcion_detallada = ?, descripcion_larga = ?, 
            variantes = ?, lista_variantes = ?, 
            stock_actual = ?,
            es_popular = ?, destacado_popular = ?,
            ocultar_web = ?
        WHERE UPPER(sku) = ?
    ''', (imagen_principal, imagenes_extra_json, imagenes_extra_json, descripcion_detallada, descripcion_detallada, merged_var_str, merged_var_str, merged_stock_total, es_popular, es_popular, ocultar_web, sku.upper()))
    
    conn.commit()
    conn.close()
    
    exportar_productos_json()
    return jsonify({
        'success': True,
        'message': f'Configuración Web de {sku} actualizada exitosamente'
    })

@app.route('/api/producto/configurar-web', methods=['POST'])
def save_producto_config_web():
    data = request.get_json() or {}
    sku = str(data.get('sku', '')).strip()
    if not sku:
        return jsonify({'error': 'SKU es requerido'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku, stock_actual, variantes, lista_variantes FROM productos WHERE UPPER(sku) = ?', (sku.upper(),))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': f'El producto con SKU "{sku}" no existe'}), 404

    imagen_principal = normalizar_nombre_imagen(str(data.get('imagen_principal', '') or data.get('imagen', '')).strip())

    raw_extras = data.get('imagenes_galeria') if data.get('imagenes_galeria') is not None else data.get('imagenes_extra')
    if raw_extras is None:
        raw_extras = data.get('imagenes', [])
        if isinstance(raw_extras, list) and len(raw_extras) > 0 and raw_extras[0] == imagen_principal:
            raw_extras = raw_extras[1:]

    if isinstance(raw_extras, list):
        imagenes_extra_json = json.dumps([img.strip() for img in raw_extras if img and img.strip()])
    else:
        imagenes_extra_json = str(raw_extras).strip()

    descripcion_detallada = str(data.get('descripcion_larga', '') or data.get('descripcion_detallada', '') or data.get('descripcion', '')).strip()
    
    existing_var_str = (row['lista_variantes'] if 'lista_variantes' in row.keys() and row['lista_variantes'] else row['variantes']) or ""
    submitted_var_str = str(data.get('lista_variantes', '') or data.get('variantes', '')).strip()
    current_stock = int(row['stock_actual']) if ('stock_actual' in row.keys() and row['stock_actual'] is not None) else 0

    merged_var_str, merged_stock_total = merge_variantes_preserving_stock(existing_var_str, submitted_var_str, current_stock)

    popular_val = data.get('destacado_popular')
    if popular_val is None:
        popular_val = data.get('es_popular')
    if popular_val is None:
        popular_val = data.get('popular')

    es_popular = 1 if popular_val in [True, 1, '1', 'true', 'True'] else 0

    ocultar_val = data.get('ocultar_web')
    if ocultar_val is None:
        ocultar_val = data.get('ocultar_en_web')
    ocultar_web = 1 if ocultar_val in [True, 1, '1', 'true', 'True'] else 0

    cursor.execute('''
        UPDATE productos
        SET imagen = ?, 
            imagenes_extra = ?, imagenes_galeria = ?, 
            descripcion_detallada = ?, descripcion_larga = ?, 
            variantes = ?, lista_variantes = ?, 
            stock_actual = ?,
            es_popular = ?, destacado_popular = ?,
            ocultar_web = ?
        WHERE UPPER(sku) = ?
    ''', (imagen_principal, imagenes_extra_json, imagenes_extra_json, descripcion_detallada, descripcion_detallada, merged_var_str, merged_var_str, merged_stock_total, es_popular, es_popular, ocultar_web, sku.upper()))

    conn.commit()
    conn.close()

    exportar_productos_json()
    return jsonify({
        'success': True,
        'message': f'Configuración Web de {sku} actualizada exitosamente'
    })

@app.route('/api/productos', methods=['POST'])
def add_producto():
    data = request.get_json() or {}
    sku = str(data.get('sku', '')).strip().upper()
    categoria = str(data.get('categoria', '')).strip()
    marca = str(data.get('marca', '')).strip()
    descripcion = str(data.get('descripcion', '')).strip()
    variantes_raw = str(data.get('variantes', '') or data.get('lista_variantes', '')).strip()
    
    try:
        costo_compra = float(data.get('costo_compra', 0))
        precio_venta = float(data.get('precio_venta', 0))
        stock_actual_input = int(data.get('stock_actual', 0))
        stock_minimo = int(data.get('stock_minimo', 2))
    except ValueError:
        return jsonify({'error': 'Valores numéricos inválidos en precios o stock'}), 400
        
    parsed_v = parse_variantes_string(variantes_raw)
    if parsed_v:
        variantes_str = ", ".join([f"{v['nombre']}:{v['stock']}" for v in parsed_v])
        if any(':' in part for part in variantes_raw.split(',')):
            stock_actual = sum(v['stock'] for v in parsed_v)
        else:
            stock_actual = stock_actual_input
    else:
        variantes_str = variantes_raw
        stock_actual = stock_actual_input

    imagen = normalizar_nombre_imagen(str(data.get('imagen', '')).strip())
    imagenes_extra = str(data.get('imagenes_extra', '')).strip()
    descripcion_detallada = str(data.get('descripcion_detallada', '')).strip()
    es_popular = 1 if data.get('es_popular') in [True, 1, '1', 'true', 'True'] else 0
    
    if not sku or not descripcion or not marca or not categoria:
        return jsonify({'error': 'El SKU, marca, categoría y descripción son obligatorios'}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku FROM productos WHERE UPPER(sku) = ?', (sku,))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': f'El SKU "{sku}" ya se encuentra registrado'}), 400
        
    cursor.execute('''
        INSERT INTO productos (sku, categoria, marca, descripcion, variantes, lista_variantes, costo_compra, precio_venta, stock_actual, stock_minimo, imagen, imagenes_extra, descripcion_detallada, es_popular)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (sku, categoria, marca, descripcion, variantes_str, variantes_str, costo_compra, precio_venta, stock_actual, stock_minimo, imagen, imagenes_extra, descripcion_detallada, es_popular))
    
    conn.commit()
    conn.close()
    
    exportar_productos_json()
    return jsonify({'message': f'Producto {sku} creado exitosamente'}), 201

@app.route('/api/productos/<sku>', methods=['PUT'])
def update_producto(sku):
    data = request.get_json() or {}
    old_sku = sku.strip().upper()
    new_sku = str(data.get('sku', old_sku)).strip().upper()
    
    categoria = str(data.get('categoria', '')).strip()
    marca = str(data.get('marca', '')).strip()
    descripcion = str(data.get('descripcion', '')).strip()
    variantes_raw = str(data.get('variantes', '') or data.get('lista_variantes', '')).strip()
    
    try:
        costo_compra = float(data.get('costo_compra', 0))
        precio_venta = float(data.get('precio_venta', 0))
        stock_actual_input = int(data.get('stock_actual', 0))
        stock_minimo = int(data.get('stock_minimo', 2))
    except ValueError:
        return jsonify({'error': 'Valores numéricos inválidos'}), 400

    if not new_sku or not descripcion or not marca or not categoria:
        return jsonify({'error': 'El SKU (código), marca, categoría y descripción son obligatorios'}), 400
        
    parsed_v = parse_variantes_string(variantes_raw)
    if parsed_v:
        variantes_str = ", ".join([f"{v['nombre']}:{v['stock']}" for v in parsed_v])
        if any(':' in part for part in variantes_raw.split(',')):
            stock_actual = sum(v['stock'] for v in parsed_v)
        else:
            stock_actual = stock_actual_input
    else:
        variantes_str = variantes_raw
        stock_actual = stock_actual_input

    imagen = normalizar_nombre_imagen(str(data.get('imagen', '')).strip())
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku FROM productos WHERE UPPER(sku) = ?', (old_sku,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': f'El producto con SKU "{old_sku}" no existe'}), 404
        
    if new_sku != old_sku:
        cursor.execute('SELECT sku FROM productos WHERE UPPER(sku) = ?', (new_sku,))
        if cursor.fetchone():
            conn.close()
            return jsonify({'error': f'El código SKU "{new_sku}" ya pertenece a otro producto registrado'}), 400

    cursor.execute('''
        UPDATE productos
        SET sku = ?, categoria = ?, marca = ?, descripcion = ?, variantes = ?, lista_variantes = ?,
            costo_compra = ?, precio_venta = ?, stock_actual = ?, stock_minimo = ?, imagen = ?
        WHERE UPPER(sku) = ?
    ''', (new_sku, categoria, marca, descripcion, variantes_str, variantes_str, costo_compra, precio_venta, stock_actual, stock_minimo, imagen, old_sku))
    
    conn.commit()
    conn.close()
    
    exportar_productos_json()
    return jsonify({'message': f'Producto {new_sku} actualizado exitosamente', 'sku': new_sku})

@app.route('/api/productos/<sku>/stock', methods=['PATCH'])
def update_stock(sku):
    data = request.get_json() or {}
    action = data.get('action', 'set')
    
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT stock_actual, variantes, lista_variantes FROM productos WHERE UPPER(sku) = ?', (sku.upper(),))
    row = cursor.fetchone()
    
    if not row:
        conn.close()
        return jsonify({'error': 'Producto no encontrado'}), 404
        
    current_stock = row['stock_actual']
    new_variantes = data.get('variantes') if data.get('variantes') is not None else data.get('lista_variantes')
    
    if new_variantes is not None:
        new_variantes_str = str(new_variantes).strip()
        if 'stock_actual' in data or 'stock_total' in data or 'new_stock' in data or ('value' in data and action != 'increment' and action != 'decrement'):
            new_stock = int(data.get('stock_actual', data.get('stock_total', data.get('new_stock', data.get('value', current_stock)))))
        else:
            parsed_v = parse_variantes_string(new_variantes_str)
            if parsed_v and any(v['stock'] > 0 for v in parsed_v):
                new_stock = sum(v['stock'] for v in parsed_v)
            else:
                new_stock = current_stock

        cursor.execute('''
            UPDATE productos 
            SET stock_actual = ?, variantes = ?, lista_variantes = ? 
            WHERE UPPER(sku) = ?
        ''', (max(0, new_stock), new_variantes_str, new_variantes_str, sku.upper()))
    else:
        try:
            value = int(data.get('value', 1))
        except (ValueError, TypeError):
            value = 1

        if action == 'increment':
            new_stock = current_stock + abs(value)
        elif action == 'decrement':
            new_stock = max(0, current_stock - abs(value))
        else:
            new_stock = max(0, value)
            
        cursor.execute('UPDATE productos SET stock_actual = ? WHERE UPPER(sku) = ?', (new_stock, sku.upper()))
        
    conn.commit()
    conn.close()
    
    exportar_productos_json()
    return jsonify({'sku': sku, 'new_stock': new_stock, 'variantes': new_variantes if new_variantes is not None else row['variantes']})

@app.route('/api/producto/ingreso-rapido', methods=['POST'])
def ingreso_rapido_producto():
    data = request.get_json() or {}
    sku = str(data.get('sku', '')).strip().upper()
    if not sku:
        return jsonify({'error': 'SKU es requerido'}), 400

    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku, stock_actual, variantes, lista_variantes FROM productos WHERE UPPER(sku) = ?', (sku,))
    row = cursor.fetchone()

    if not row:
        conn.close()
        return jsonify({'error': f'El producto con SKU "{sku}" no existe'}), 404

    current_stock = int(row['stock_actual'])
    var_str = (row['lista_variantes'] if 'lista_variantes' in row.keys() and row['lista_variantes'] else row['variantes']) or ""
    
    ingresos = data.get('ingresos')
    
    if isinstance(ingresos, dict) and len(ingresos) > 0:
        parsed_variants = parse_variantes_string(var_str)
        variants_dict = {}
        for v in parsed_variants:
            variants_dict[v['nombre']] = v['stock']

        for v_name, qty_add in ingresos.items():
            v_name_clean = str(v_name).strip()
            try:
                added = int(qty_add)
            except (ValueError, TypeError):
                added = 0
            current_v_stock = variants_dict.get(v_name_clean, 0)
            variants_dict[v_name_clean] = max(0, current_v_stock + max(0, added))

        new_variants_parts = [f"{name}:{st}" for name, st in variants_dict.items()]
        new_variants_str = ", ".join(new_variants_parts)
        new_total_stock = sum(variants_dict.values())

        cursor.execute('''
            UPDATE productos 
            SET stock_actual = ?, variantes = ?, lista_variantes = ? 
            WHERE UPPER(sku) = ?
        ''', (new_total_stock, new_variants_str, new_variants_str, sku))
    else:
        try:
            qty_add = int(data.get('cantidad', data.get('ingresos', 0)))
        except (ValueError, TypeError):
            qty_add = 0
            
        if qty_add <= 0:
            conn.close()
            return jsonify({'error': 'La cantidad a ingresar debe ser mayor a 0'}), 400

        new_total_stock = current_stock + qty_add
        new_variants_str = var_str

        cursor.execute('''
            UPDATE productos 
            SET stock_actual = ?
            WHERE UPPER(sku) = ?
        ''', (new_total_stock, sku))

    conn.commit()
    conn.close()

    exportar_productos_json()

    return jsonify({
        'success': True,
        'message': f'Ingreso de mercadería registrado para {sku}',
        'sku': sku,
        'stock_actual': new_total_stock,
        'variantes': new_variants_str
    })

@app.route('/api/productos/<sku>', methods=['DELETE'])
def delete_producto(sku):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT sku FROM productos WHERE sku = ?', (sku,))
    if not cursor.fetchone():
        conn.close()
        return jsonify({'error': 'Producto no encontrado'}), 404
        
    cursor.execute('DELETE FROM productos WHERE sku = ?', (sku,))
    conn.commit()
    conn.close()
    
    exportar_productos_json()
    return jsonify({'message': f'Producto {sku} eliminado exitosamente'})

@app.route('/api/exportar-excel', methods=['GET'])
def exportar_excel():
    """Genera y descarga un archivo nativo .xlsx con columnas individuales formateadas con openpyxl."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT sku, categoria, marca, descripcion, variantes, costo_compra, precio_venta, stock_actual, stock_minimo
            FROM productos
            ORDER BY categoria, marca, descripcion
        ''')
        rows = cursor.fetchall()
        conn.close()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventario AGR"
        ws.views.sheetView[0].showGridLines = True

        headers = [
            'SKU', 'Categoría', 'Marca', 'Descripción / Modelo', 'Variantes',
            'Costo Compra (S/)', 'Precio Venta (S/)', 'Stock Actual', 'Stock Mínimo', 'Estado'
        ]
        
        ws.append(headers)

        header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        ws.row_dimensions[1].height = 25

        for r in rows:
            stock = int(r['stock_actual'])
            stock_min = int(r['stock_minimo'])
            
            if stock == 0:
                estado = "Agotado"
            elif stock <= stock_min:
                estado = "Reponer"
            else:
                estado = "En Stock"

            row_data = [
                r['sku'],
                r['categoria'],
                r['marca'],
                r['descripcion'],
                r['variantes'] or '',
                float(r['costo_compra']),
                float(r['precio_venta']),
                stock,
                stock_min,
                estado
            ]
            ws.append(row_data)

            current_row = ws.max_row
            ws.cell(row=current_row, column=1).alignment = center_align
            ws.cell(row=current_row, column=2).alignment = left_align
            ws.cell(row=current_row, column=3).alignment = left_align
            ws.cell(row=current_row, column=4).alignment = left_align
            ws.cell(row=current_row, column=5).alignment = left_align
            ws.cell(row=current_row, column=6).number_format = '"S/" #,##0.00'
            ws.cell(row=current_row, column=6).alignment = right_align
            ws.cell(row=current_row, column=7).number_format = '"S/" #,##0.00'
            ws.cell(row=current_row, column=7).alignment = right_align
            ws.cell(row=current_row, column=8).alignment = center_align
            ws.cell(row=current_row, column=9).alignment = center_align
            ws.cell(row=current_row, column=10).alignment = center_align

        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"inventario_agr_importaciones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        print(f"Error exportando Excel: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/importar-excel', methods=['POST'])
def importar_excel():
    """Procesa e importa un archivo .xlsx o .csv aplicando UPSERT en la BD y actualizando public/productos.json."""
    if 'file' not in request.files:
        return jsonify({'error': 'No se proporcionó ningún archivo'}), 400

    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'error': 'Archivo no seleccionado'}), 400

    filename = file.filename.lower()
    items_to_process = []

    try:
        if filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                return jsonify({'error': 'El archivo Excel no contiene filas de datos válidas'}), 400

            header_raw = [str(cell).strip().lower() if cell is not None else '' for cell in rows[0]]
            
            sku_idx = next((i for i, h in enumerate(header_raw) if 'sku' in h or 'código' in h or 'codigo' in h), 0)
            cat_idx = next((i for i, h in enumerate(header_raw) if 'categor' in h), 1)
            marca_idx = next((i for i, h in enumerate(header_raw) if 'marca' in h), 2)
            desc_idx = next((i for i, h in enumerate(header_raw) if 'descrip' in h or 'producto' in h or 'modelo' in h), 3)
            var_idx = next((i for i, h in enumerate(header_raw) if 'variant' in h), 4)
            costo_idx = next((i for i, h in enumerate(header_raw) if 'costo' in h), 5)
            precio_idx = next((i for i, h in enumerate(header_raw) if 'precio' in h or 'venta' in h), 6)
            stock_idx = next((i for i, h in enumerate(header_raw) if 'stock actual' in h or 'stock' in h), 7)
            min_idx = next((i for i, h in enumerate(header_raw) if 'mínimo' in h or 'minimo' in h), 8)

            for row in rows[1:]:
                if not row or not any(row):
                    continue
                sku = str(row[sku_idx]).strip().upper() if len(row) > sku_idx and row[sku_idx] is not None else ''
                if not sku or sku == 'NONE':
                    continue

                def safe_str(val, default=''):
                    return str(val).strip() if val is not None else default

                def clean_currency_float(val, default=0.0):
                    if val is None or val == '':
                        return default
                    if isinstance(val, (int, float)):
                        return float(val)
                    s = str(val).replace('S/', '').replace('$', '').replace('s/', '').replace('S/.', '').replace('s/.', '').replace(' ', '').replace(',', '.').strip()
                    try:
                        return float(s)
                    except Exception:
                        cleaned = re.sub(r'[^0-9.]', '', s)
                        try:
                            return float(cleaned) if cleaned else default
                        except Exception:
                            return default

                def clean_int(val, default=0):
                    if val is None or val == '':
                        return default
                    if isinstance(val, int):
                        return val
                    if isinstance(val, float):
                        return int(val)
                    try:
                        s = str(val).strip()
                        return int(float(s))
                    except Exception:
                        cleaned = re.sub(r'[^0-9]', '', str(val))
                        return int(cleaned) if cleaned else default

                categoria = safe_str(row[cat_idx] if len(row) > cat_idx else '', 'GENERAL')
                marca = safe_str(row[marca_idx] if len(row) > marca_idx else '', 'AGR')
                descripcion = safe_str(row[desc_idx] if len(row) > desc_idx else '', sku)
                variantes = safe_str(row[var_idx] if len(row) > var_idx else '', '')
                costo_compra = clean_currency_float(row[costo_idx] if len(row) > costo_idx else 0.0)
                precio_venta = clean_currency_float(row[precio_idx] if len(row) > precio_idx else 0.0)
                stock_actual = clean_int(row[stock_idx] if len(row) > stock_idx else 0)
                stock_minimo = clean_int(row[min_idx] if len(row) > min_idx else 2)

                items_to_process.append((sku, categoria, marca, descripcion, variantes, costo_compra, precio_venta, stock_actual, stock_minimo))

        elif filename.endswith('.csv'):
            stream = file.stream.read().decode('utf-8-sig', errors='replace')
            lines = stream.splitlines()
            if not lines:
                return jsonify({'error': 'El archivo CSV está vacío'}), 400

            delimiter = ';' if ';' in lines[0] else ','
            reader = csv.reader(lines, delimiter=delimiter)
            rows = list(reader)
            if len(rows) < 2:
                return jsonify({'error': 'El archivo CSV no contiene filas de datos'}), 400

            header_raw = [h.strip().lower() for h in rows[0]]
            sku_idx = next((i for i, h in enumerate(header_raw) if 'sku' in h or 'código' in h or 'codigo' in h), 0)
            cat_idx = next((i for i, h in enumerate(header_raw) if 'categor' in h), 1)
            marca_idx = next((i for i, h in enumerate(header_raw) if 'marca' in h), 2)
            desc_idx = next((i for i, h in enumerate(header_raw) if 'descrip' in h or 'producto' in h or 'modelo' in h), 3)
            var_idx = next((i for i, h in enumerate(header_raw) if 'variant' in h), 4)
            costo_idx = next((i for i, h in enumerate(header_raw) if 'costo' in h), 5)
            precio_idx = next((i for i, h in enumerate(header_raw) if 'precio' in h or 'venta' in h), 6)
            stock_idx = next((i for i, h in enumerate(header_raw) if 'stock actual' in h or 'stock' in h), 7)
            min_idx = next((i for i, h in enumerate(header_raw) if 'mínimo' in h or 'minimo' in h), 8)

            for row in rows[1:]:
                if not row or not any(row):
                    continue
                sku = row[sku_idx].strip().upper() if len(row) > sku_idx else ''
                if not sku:
                    continue

                def clean_currency_float(val, default=0.0):
                    if val is None or val == '':
                        return default
                    if isinstance(val, (int, float)):
                        return float(val)
                    s = str(val).replace('S/', '').replace('$', '').replace('s/', '').replace('S/.', '').replace('s/.', '').replace(' ', '').replace(',', '.').strip()
                    try:
                        return float(s)
                    except Exception:
                        cleaned = re.sub(r'[^0-9.]', '', s)
                        try:
                            return float(cleaned) if cleaned else default
                        except Exception:
                            return default

                def clean_int(val, default=0):
                    if val is None or val == '':
                        return default
                    if isinstance(val, int):
                        return val
                    if isinstance(val, float):
                        return int(val)
                    try:
                        s = str(val).strip()
                        return int(float(s))
                    except Exception:
                        cleaned = re.sub(r'[^0-9]', '', str(val))
                        return int(cleaned) if cleaned else default

                categoria = row[cat_idx].strip() if len(row) > cat_idx and row[cat_idx] else 'GENERAL'
                marca = row[marca_idx].strip() if len(row) > marca_idx and row[marca_idx] else 'AGR'
                descripcion = row[desc_idx].strip() if len(row) > desc_idx and row[desc_idx] else sku
                variantes = row[var_idx].strip() if len(row) > var_idx else ''
                costo_compra = clean_currency_float(row[costo_idx] if len(row) > costo_idx else 0.0)
                precio_venta = clean_currency_float(row[precio_idx] if len(row) > precio_idx else 0.0)
                stock_actual = clean_int(row[stock_idx] if len(row) > stock_idx else 0)
                stock_minimo = clean_int(row[min_idx] if len(row) > min_idx else 2)

                items_to_process.append((sku, categoria, marca, descripcion, variantes, costo_compra, precio_venta, stock_actual, stock_minimo))

        else:
            return jsonify({'error': 'Formato no soportado. Sube un archivo .xlsx o .csv'}), 400

        if not items_to_process:
            return jsonify({'error': 'No se encontraron productos válidos en el archivo'}), 400

        conn = get_db()
        cursor = conn.cursor()
        inserted_count = 0
        updated_count = 0

        for item in items_to_process:
            sku, categoria, marca, descripcion, variantes, costo_compra, precio_venta, stock_actual, stock_minimo = item
            cursor.execute('SELECT sku FROM productos WHERE UPPER(sku) = ?', (sku.upper(),))
            exists = cursor.fetchone()

            if exists:
                cursor.execute('''
                    UPDATE productos
                    SET categoria = ?, marca = ?, descripcion = ?, variantes = ?, lista_variantes = ?,
                        costo_compra = ?, precio_venta = ?, stock_actual = ?, stock_minimo = ?
                    WHERE UPPER(sku) = ?
                ''', (categoria, marca, descripcion, variantes, variantes, costo_compra, precio_venta, stock_actual, stock_minimo, sku.upper()))
                updated_count += 1
            else:
                cursor.execute('''
                    INSERT INTO productos (sku, categoria, marca, descripcion, variantes, lista_variantes, costo_compra, precio_venta, stock_actual, stock_minimo, imagen)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ''', (sku, categoria, marca, descripcion, variantes, variantes, costo_compra, precio_venta, stock_actual, stock_minimo, f"{sku}.jpg"))
                inserted_count += 1

        conn.commit()
        conn.close()

        exportar_productos_json()

        total_processed = inserted_count + updated_count
        return jsonify({
            'success': True,
            'inserted': inserted_count,
            'updated': updated_count,
            'total': total_processed,
            'message': f'¡Importación completada: {total_processed} productos procesados exitosamente! ({inserted_count} creados, {updated_count} actualizados)'
        })

    except Exception as e:
        print(f"Error importando Excel: {e}")
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500

@app.route('/api/parse-whatsapp', methods=['POST'])
def parse_whatsapp_order():
    data = request.get_json() or {}
    raw_text = data.get('text', '').strip()
    
    if not raw_text:
        return jsonify({'error': 'Texto de WhatsApp vacío'}), 400
        
    items_to_find = []
    
    # 1. Patron explicito: CODIGO_PEDIDO:[SX20:5,MQ-20:2] o [SX20:5, MQ-20:2]
    match_explicit = re.search(r'\[([^\]]+)\]', raw_text)
    if match_explicit:
        content = match_explicit.group(1)
        pairs = content.split(',')
        for pair in pairs:
            if ':' in pair:
                parts = pair.split(':')
                sku_code = parts[0].strip().upper()
                try:
                    qty = int(parts[1].strip())
                    items_to_find.append((sku_code, qty))
                except ValueError:
                    pass
    
    # 2. Si no se halló patrón explícito o para complementar, usar expresiones regulares sobre líneas/palabras
    if not items_to_find:
        # Formatos comunes: SKU:CANT, SKU x CANT, SKU - CANT, CANT de SKU, SKU CANT
        # Ejemplo: SX20: 5, SX20 x5, 5x SX20, SX20 5
        patterns = [
            r'([A-Za-z0-9\-]{2,15})[\s:]*[xX*:-]\s*(\d+)', # SX20:5 o SX20 x 5 o SX20-5
            r'(\d+)\s*[xX*:-]?\s*([A-Za-z0-9\-]{2,15})',   # 5 x SX20 o 5 SX20
            r'([A-Za-z0-9\-]{2,15})\s+(\d+)'                # SX20 5
        ]
        
        seen_skus = set()
        for line in raw_text.split('\n'):
            line = line.strip()
            if not line:
                continue
            matched_line = False
            for p in patterns:
                matches = re.findall(p, line)
                if matches:
                    for m1, m2 in matches:
                        # Verificar cual es el SKU y cual es la cantidad
                        if m1.isdigit():
                            qty = int(m1)
                            sku_code = m2.strip().upper()
                        elif m2.isdigit():
                            sku_code = m1.strip().upper()
                            qty = int(m2)
                        else:
                            continue
                            
                        if len(sku_code) >= 2 and sku_code not in seen_skus:
                            items_to_find.append((sku_code, qty))
                            seen_skus.add(sku_code)
                            matched_line = True
                            break
                if matched_line:
                    break

    if not items_to_find:
        return jsonify({
            'success': False,
            'message': 'No se detectaron códigos SKU o cantidades en el mensaje.',
            'items': []
        }), 400

    conn = get_db()
    cursor = conn.cursor()
    
    parsed_items = []
    unfound_skus = []
    out_of_stock_skus = []
    
    for sku_code, qty_requested in items_to_find:
        cursor.execute('SELECT * FROM productos WHERE UPPER(sku) = ?', (sku_code,))
        row = cursor.fetchone()
        if row:
            stock = int(row['stock_actual'])
            precio = float(row['precio_venta'])
            is_stock_ok = stock >= qty_requested
            
            if not is_stock_ok:
                out_of_stock_skus.append(f"{sku_code} (Stock: {stock}, Solicitado: {qty_requested})")
                
            parsed_items.append({
                'sku': row['sku'],
                'categoria': row['categoria'],
                'marca': row['marca'],
                'descripcion': row['descripcion'],
                'variantes': row['variantes'] or '',
                'precio_venta': precio,
                'cantidad': qty_requested,
                'stock_actual': stock,
                'stock_suficiente': is_stock_ok,
                'subtotal': round(precio * qty_requested, 2),
                'imagen': row['imagen'] or ''
            })
        else:
            unfound_skus.append(sku_code)
            
    conn.close()
    
    return jsonify({
        'success': True,
        'parsed_items': parsed_items,
        'unfound_skus': unfound_skus,
        'out_of_stock_skus': out_of_stock_skus,
        'total_items': len(parsed_items)
    })

@app.route('/api/procesar-venta', methods=['POST'])
@app.route('/api/ventas', methods=['POST'])
def process_sale():
    data = request.get_json() or {}
    cliente = str(data.get('cliente', 'Cliente General')).strip() or 'Cliente General'
    items = data.get('items', [])
    
    if not items or not isinstance(items, list):
        return jsonify({'error': 'La venta debe contener al menos un producto'}), 400
        
    conn = get_db()
    cursor = conn.cursor()
    
    validated_items = []
    total_venta = 0.0
    
    try:
        # Descontar stock por variante e ítem
        for item in items:
            sku = str(item.get('sku', '')).strip().upper()
            variant_target = str(item.get('variante', item.get('variantes', 'Único'))).strip()
            qty = int(item.get('cantidad', 1))
            if qty <= 0:
                continue
                
            cursor.execute('SELECT * FROM productos WHERE UPPER(sku) = ?', (sku,))
            row = cursor.fetchone()
            if not row:
                conn.close()
                return jsonify({'error': f'El producto con SKU "{sku}" no existe en el sistema'}), 400
                
            costo = float(row['costo_compra']) if ('costo_compra' in row.keys() and row['costo_compra'] is not None) else 0.0
            precio_unitario = float(item.get('precio', item.get('precio_venta', row['precio_venta'])))
            
            raw_var_str = row['variantes'] or row['lista_variantes'] or ''
            parsed_variants = parse_variantes_string(raw_var_str)
            
            new_variants_str = raw_var_str
            new_stock_actual = int(row['stock_actual'])
            
            # Buscar variante específica en la lista parsed
            matching_var = None
            if variant_target and variant_target.upper() not in ['ÚNICO', 'UNICO', 'ESTÁNDAR', 'ESTANDAR']:
                for v in parsed_variants:
                    if v['nombre'].strip().upper() == variant_target.upper():
                        matching_var = v
                        break
            
            if matching_var:
                if matching_var['stock'] < qty:
                    conn.close()
                    return jsonify({'error': f'Stock insuficiente para variante "{matching_var["nombre"]}" de {row["descripcion"]} ({sku}). Stock disponible: {matching_var["stock"]}, Solicitado: {qty}'}), 400
                
                matching_var['stock'] -= qty
                new_variants_parts = [f"{v['nombre']}:{v['stock']}" for v in parsed_variants]
                new_variants_str = ", ".join(new_variants_parts)
                new_stock_actual = sum(v['stock'] for v in parsed_variants)
                
                cursor.execute('''
                    UPDATE productos 
                    SET stock_actual = ?, variantes = ?, lista_variantes = ? 
                    WHERE UPPER(sku) = ?
                ''', (new_stock_actual, new_variants_str, new_variants_str, sku))
            else:
                if new_stock_actual < qty:
                    conn.close()
                    return jsonify({'error': f'Stock insuficiente para {row["descripcion"]} ({sku}). Stock disponible: {new_stock_actual}, Solicitado: {qty}'}), 400
                
                new_stock_actual = max(0, new_stock_actual - qty)
                cursor.execute('''
                    UPDATE productos 
                    SET stock_actual = ? 
                    WHERE UPPER(sku) = ?
                ''', (new_stock_actual, sku))

            subtotal = round(precio_unitario * qty, 2)
            ganancia_item = round((precio_unitario - costo) * qty, 2)
            total_venta += subtotal
            
            validated_items.append({
                'sku': row['sku'],
                'marca': row['marca'],
                'descripcion': row['descripcion'],
                'variante': variant_target if variant_target else 'Único',
                'variantes': variant_target if variant_target else 'Único',
                'costo_compra': costo,
                'precio_unitario': precio_unitario,
                'cantidad': qty,
                'subtotal': subtotal,
                'ganancia': ganancia_item
            })

        if not validated_items:
            conn.close()
            return jsonify({'error': 'No hay items válidos para procesar'}), 400
            
        fecha_hora = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        detalle_json = json.dumps(validated_items, ensure_ascii=False)
        
        cursor.execute('''
            INSERT INTO ventas (fecha_hora, cliente, total, detalle_json)
            VALUES (?, ?, ?, ?)
        ''', (fecha_hora, cliente, round(total_venta, 2), detalle_json))
        
        venta_id = cursor.lastrowid
        conn.commit()
        conn.close()
        
        exportar_productos_json()
        
        return jsonify({
            'status': 'success',
            'success': True,
            'message': 'Venta procesada exitosamente',
            'venta_id': venta_id,
            'cliente': cliente,
            'total': round(total_venta, 2),
            'fecha_hora': fecha_hora,
            'items_count': len(validated_items)
        }), 200
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Error procesando la venta: {str(e)}'}), 500

@app.route('/api/ventas/<int:venta_id>', methods=['DELETE'])
def delete_venta(venta_id):
    """Anula una venta, devuelve los productos al stock del almacén y elimina el registro de venta."""
    try:
        conn = get_db()
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM ventas WHERE id = ?', (venta_id,))
        venta = cursor.fetchone()
        
        if not venta:
            conn.close()
            return jsonify({'error': f'La venta #{venta_id} no fue encontrada.'}), 404
            
        try:
            items = json.loads(venta['detalle_json'])
        except Exception:
            items = []
            
        for item in items:
            sku = item.get('sku')
            qty = int(item.get('cantidad', 0))
            variant_name = str(item.get('variante', item.get('variantes', ''))).strip()
            if sku and qty > 0:
                cursor.execute('SELECT * FROM productos WHERE UPPER(sku) = UPPER(?)', (sku,))
                p_row = cursor.fetchone()
                if p_row:
                    raw_var_str = p_row['variantes'] or p_row['lista_variantes'] or ''
                    parsed_vars = parse_variantes_string(raw_var_str)
                    matching_v = None
                    if variant_name and variant_name.upper() not in ['ÚNICO', 'UNICO', 'ESTÁNDAR', 'ESTANDAR']:
                        for v in parsed_vars:
                            if v['nombre'].strip().upper() == variant_name.upper():
                                matching_v = v
                                break
                    if matching_v:
                        matching_v['stock'] += qty
                        new_var_parts = [f"{v['nombre']}:{v['stock']}" for v in parsed_vars]
                        new_var_str = ", ".join(new_var_parts)
                        new_tot_stock = sum(v['stock'] for v in parsed_vars)
                        cursor.execute('UPDATE productos SET stock_actual = ?, variantes = ?, lista_variantes = ? WHERE UPPER(sku) = UPPER(?)', (new_tot_stock, new_var_str, new_var_str, sku))
                    else:
                        cursor.execute('UPDATE productos SET stock_actual = stock_actual + ? WHERE UPPER(sku) = UPPER(?)', (qty, sku))
                
        cursor.execute('DELETE FROM ventas WHERE id = ?', (venta_id,))
        conn.commit()
        conn.close()
        
        exportar_productos_json()
        
        return jsonify({'message': f'Venta #{venta_id} anulada exitosamente y stock devuelto a almacén.'})
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': f'Error anulando la venta: {str(e)}'}), 500

@app.route('/api/reportes', methods=['GET'])
def get_reportes():
    search_q = request.args.get('q', '').strip().lower()
    desde_str = request.args.get('desde', '').strip()
    hasta_str = request.args.get('hasta', '').strip()
    
    conn = get_db()
    cursor = conn.cursor()

    # Cargar costos actuales de productos para cálculo retroactivo de ganancias
    cursor.execute('SELECT sku, costo_compra FROM productos')
    costos_map = {r['sku'].upper(): float(r['costo_compra']) for r in cursor.fetchall()}
    
    # 1. Valor total del inventario
    cursor.execute('''
        SELECT 
            COALESCE(SUM(stock_actual * costo_compra), 0) as valor_costo,
            COALESCE(SUM(stock_actual * precio_venta), 0) as valor_venta,
            COUNT(*) as total_productos,
            SUM(CASE WHEN stock_actual = 0 THEN 1 ELSE 0 END) as productos_agotados,
            SUM(CASE WHEN stock_actual > 0 AND stock_actual <= stock_minimo THEN 1 ELSE 0 END) as productos_reponer
        FROM productos
    ''')
    row_inv = cursor.fetchone()
    valor_costo = float(row_inv['valor_costo'])
    valor_venta = float(row_inv['valor_venta'])
    ganancia_potencial = valor_venta - valor_costo
    total_productos = int(row_inv['total_productos'])
    productos_agotados = int(row_inv['productos_agotados'] or 0)
    productos_reponer = int(row_inv['productos_reponer'] or 0)

    # 2. Construir SQL de ventas con filtros opcionales de fecha y buscador
    sql_ventas = 'SELECT * FROM ventas WHERE 1=1'
    params = []

    if desde_str:
        sql_ventas += ' AND fecha_hora >= ?'
        params.append(f'{desde_str} 00:00:00')
    if hasta_str:
        sql_ventas += ' AND fecha_hora <= ?'
        params.append(f'{hasta_str} 23:59:59')

    sql_ventas += ' ORDER BY id DESC'

    cursor.execute(sql_ventas, params)
    rows_ventas = cursor.fetchall()

    historial_ventas = []
    recaudado_total = 0.0
    ganancia_neta_total = 0.0

    for r in rows_ventas:
        v_id = r['id']
        f_hora = r['fecha_hora']
        cliente = r['cliente'] or 'Cliente General'
        total_v = float(r['total'])

        try:
            items = json.loads(r['detalle_json'])
        except Exception:
            items = []

        total_cant_items = 0
        ganancia_ticket = 0.0

        items_processed = []
        for i in items:
            sku = str(i.get('sku', '')).upper()
            cant = int(i.get('cantidad', 1))
            precio_u = float(i.get('precio_unitario', i.get('precio_venta', 0)))
            costo_u = float(i.get('costo_compra', costos_map.get(sku, 0)))
            subtotal_item = round(precio_u * cant, 2)
            ganancia_item = round((precio_u - costo_u) * cant, 2)

            total_cant_items += cant
            ganancia_ticket += ganancia_item

            items_processed.append({
                'sku': sku,
                'descripcion': i.get('descripcion', ''),
                'marca': i.get('marca', ''),
                'variantes': i.get('variantes', ''),
                'cantidad': cant,
                'costo_compra': round(costo_u, 2),
                'precio_unitario': round(precio_u, 2),
                'subtotal': subtotal_item,
                'ganancia': ganancia_item
            })

        # Aplicar filtro de búsqueda de texto si fue ingresado
        if search_q:
            match_id = str(v_id) in search_q
            match_client = search_q in cliente.lower()
            match_prod = any(search_q in it['sku'].lower() or search_q in it['descripcion'].lower() for it in items_processed)
            if not (match_id or match_client or match_prod):
                continue

        recaudado_total += total_v
        ganancia_neta_total += ganancia_ticket

        historial_ventas.append({
            'id': v_id,
            'fecha_hora': f_hora,
            'cliente': cliente,
            'total': round(total_v, 2),
            'ganancia': round(ganancia_ticket, 2),
            'total_cant': total_cant_items,
            'items': items_processed
        })

    # 3. Ventas del día de hoy (para métricas rápidas)
    hoy_str = datetime.now().strftime('%Y-%m-%d')
    cursor.execute("SELECT COALESCE(SUM(total), 0) as total_dia, COUNT(*) as count_dia FROM ventas WHERE fecha_hora LIKE ?", (f'{hoy_str}%',))
    row_dia = cursor.fetchone()
    ventas_dia = float(row_dia['total_dia'])
    transacciones_dia = int(row_dia['count_dia'])

    conn.close()

    return jsonify({
        'recaudado_total': round(recaudado_total, 2),
        'ganancia_neta': round(ganancia_neta_total, 2),
        'ventas_dia': round(ventas_dia, 2),
        'transacciones_dia': transacciones_dia,
        'valor_costo': round(valor_costo, 2),
        'valor_venta': round(valor_venta, 2),
        'ganancia_potencial': round(ganancia_potencial, 2),
        'total_productos': total_productos,
        'productos_agotados': productos_agotados,
        'productos_reponer': productos_reponer,
        'historial_ventas': historial_ventas
    })

@app.route('/api/subir-imagen', methods=['POST'])
@app.route('/api/upload-imagen', methods=['POST'])
def upload_imagen():
    file = request.files.get('file') or request.files.get('imagen')
    if not file or file.filename == '':
        return jsonify({'status': 'error', 'error': 'No file selected'}), 400
        
    filename = secure_filename(file.filename)
    if not filename:
        filename = f"img_{int(datetime.now().timestamp())}.jpg"

    base, ext = os.path.splitext(filename)
    ext = ext.lower()
    allowed_exts = ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.svg']
    if ext not in allowed_exts:
        return jsonify({'status': 'error', 'error': 'Formato de imagen no permitido (.jpg, .jpeg, .png, .webp)'}), 400
        
    filename = f"{base}{ext}"
    save_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(save_path)

    # Disparar auto-sincronización con GitHub y Vercel en segundo plano
    sincronizar_con_github()

    return jsonify({
        'status': 'success',
        'message': 'Imagen subida correctamente',
        'filename': filename,
        'url': f'imagenes/{filename}'
    }), 200

@app.route('/api/sincronizar', methods=['POST'])
def sync_catalog():
    success = exportar_productos_json()
    if success:
        return jsonify({'message': 'Catálogo web sincronizado con éxito en public/productos.json'})
    else:
        return jsonify({'error': 'Fallo al sincronizar productos.json'}), 500

if __name__ == '__main__':
    print("Iniciando Sistema de Gestión de Inventario AGR Importaciones en http://127.0.0.1:5000")
    app.run(host='0.0.0.0', port=5000, debug=True)
